# -*- coding: utf-8 -*-
"""
掃描「學生作業」下各資料夾的「{資料夾名}_評分表.csv」，
彙整成一個 Excel 總表：每個學生一個工作表，工作表名稱 = 資料夾名稱。
輸出檔：Data/評分表_總表.xlsx（掃描時排除 Script/Data/Doc/Gallery/__pycache__）
"""
import csv
import os
import re

try:
    from paths_config import BASE_DIR, DATA_DIR, iter_student_folders
except ImportError:
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    DATA_DIR = os.path.join(BASE_DIR, "Data")
    def iter_student_folders():
        skip = {"Script", "Data", "Doc", "Gallery", "__pycache__"}
        for name in sorted(os.listdir(BASE_DIR)):
            path = os.path.join(BASE_DIR, name)
            if os.path.isdir(path) and name not in skip:
                yield name


def safe_sheet_name(name, max_len=31):
    """Excel 工作表名稱不可含 \\ / * ? [ ] :，且最長 31 字元。"""
    s = re.sub(r'[\\/*?:\[\]]', '_', name)
    return s[:max_len] if len(s) > max_len else s


def load_csv_rows(csv_path):
    """用標準 csv 模組讀取，支援欄位內逗號與引號。"""
    with open(csv_path, "r", encoding="utf-8") as f:
        return list(csv.reader(f))


def write_csv_rows(csv_path, rows):
    """將 rows 寫回 CSV（含逗號的欄位會自動加引號）。"""
    with open(csv_path, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
        for row in rows:
            w.writerow(row)


def star_count_to_score(s):
    """將「得分」欄的 ★★★★★ / ★★★★ 等轉成 1~5（數星星個數）。"""
    s = (s or "").strip()
    n = len(re.findall(r"★", s))
    return max(1, min(5, n)) if n else 0


def rows_with_avg_score_header(rows):
    """
    將標題列的「得分 (5★)」改為該生五維星等的平均，例如「得分 (3.8★)」。
    若找不到得分欄或不足 5 筆資料，原樣回傳。
    """
    if not rows or len(rows) < 6:
        return rows
    header = list(rows[0])
    col_idx = None
    for i, cell in enumerate(header):
        if re.match(r"得分\s*\(.*★", str(cell).strip()):
            col_idx = i
            break
    if col_idx is None:
        return rows
    scores = []
    for row in rows[1:6]:
        if len(row) > col_idx:
            scores.append(star_count_to_score(row[col_idx]))
    if len(scores) != 5:
        return rows
    avg = round(sum(scores) / 5, 1)
    header[col_idx] = f"得分 ({avg}★)"
    return [header] + rows[1:]


def build_total_xlsx(base_dir=None, output_filename="評分表_總表.xlsx"):
    base_dir = base_dir or BASE_DIR
    output_path = os.path.join(DATA_DIR, output_filename)

    try:
        from openpyxl import Workbook
    except ImportError:
        print("請先安裝： pip install openpyxl")
        return

    os.makedirs(DATA_DIR, exist_ok=True)
    # 學生資料夾（排除 Script/Data/Doc/Gallery/__pycache__）
    all_students = list(iter_student_folders())
    success_list = []
    skip_list = []

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    # 先收集所有成功讀取的 (名稱, rows)
    collected = []
    for name in all_students:
        folder = os.path.join(base_dir, name)
        csv_path = os.path.join(folder, f"{name}_評分表.csv")
        if not os.path.isfile(csv_path):
            skip_list.append(name)
            continue
        rows = load_csv_rows(csv_path)
        if not rows:
            skip_list.append(name)
            continue
        collected.append((name, rows))
        success_list.append(name)
        print(f"  已彙整: {name}")

    # 將各學生 CSV 的標題改為「得分 (平均★)」並寫回
    for name, rows in collected:
        rows_show = rows_with_avg_score_header(rows)
        csv_path = os.path.join(BASE_DIR, name, f"{name}_評分表.csv")
        write_csv_rows(csv_path, rows_show)

    # 最前面開一個「總覽」sheet：所有人的資料貼在一起，每人前加姓名、中間空一行；標題改為該生平均星等
    if collected:
        overview = wb.create_sheet(title="總覽", index=0)
        for i, (name, rows) in enumerate(collected):
            rows_show = rows_with_avg_score_header(rows)
            num_cols = len(rows_show[0]) if rows_show else 1
            name_row = [name] + [""] * (num_cols - 1)
            overview.append(name_row)
            for row in rows_show:
                overview.append(row)
            if i < len(collected) - 1:
                overview.append([])  # 空一行

    # 各學生獨立 sheet（在總覽之後）；標題改為該生平均星等
    for name, rows in collected:
        rows_show = rows_with_avg_score_header(rows)
        sheet_title = safe_sheet_name(name)
        ws = wb.create_sheet(title=sheet_title)
        for row in rows_show:
            ws.append(row)

    if success_list:
        wb.save(output_path)
        print()
        print("---------- 彙整結果 ----------")
        print(f"已產生總表：{output_path}")
        print(f"成功彙整（共 {len(success_list)} 位）：")
        for s in success_list:
            print(f"  · {s}")
        if skip_list:
            print(f"跳過（缺少評分所需檔案，共 {len(skip_list)} 位）：")
            for s in skip_list:
                print(f"  · {s}")
    else:
        print("未找到任何「{資料夾名}_評分表.csv」，未產生總表。")
        if skip_list:
            print(f"跳過（缺少評分所需檔案，共 {len(skip_list)} 位）：")
            for s in skip_list:
                print(f"  · {s}")


if __name__ == "__main__":
    build_total_xlsx()
